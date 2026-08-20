import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")

path = r"C:\Users\aaa\coding\dashboard\data\재무PPT_이상항목_종합보고서.xlsx"
wb = load_workbook(path)

SHEET_NAME = "05_파일명상이항목"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)

headers = ["파일 위치", "표준 형식과 다른 점", "비고"]
widths = [78, 40, 45]

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
c = ws.cell(1, 1, "파일명 표준 형식 상이 항목")
c.font = TITLE_FONT
ws.row_dimensions[1].height = 22
for i, h in enumerate(headers, start=1):
    cell = ws.cell(2, i, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"
ws.row_dimensions[2].height = 24

rows = [
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\신사업기획파트\[경기테크노파크] 친환경차 부품개발 인력양성 위탁교육 착수보고_260513.pptx",
     "파트명(_신사업_) 없음, 연도(26년) 텍스트 없음, 단계 표기가 '_착수보고_260513'(날짜코드)로 표준([단계]/_단계)과 다름",
     ""),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\전동화&차량개발교육파트\(HMC RnD 개별) 26년_해외연구소 실무역량 강화_제안.pptx",
     "파트명(_전차_) 누락",
     "같은 폴더에 정상본 '(HMC RnD) 26년_해외연구소 실무역량 강화_전차_제안.pptx' 존재 — 사실상 중복 파일로 추정, 담당자 확인 필요"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\전동화&차량개발교육파트\(HMC RnD 개별) 26년_해외연구소 실무역량 강화_착수.pptx",
     "파트명(_전차_) 누락",
     "같은 폴더에 정상본 '(HMC RnD) 26년_해외연구소 실무역량 강화_전차_착수.pptx' 존재 — 사실상 중복 파일로 추정, 담당자 확인 필요"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\전동화&차량개발교육파트\(HMC) 26년_R&DAVP본부 外 (ICT GSO 등) 특허 교육 운영_전차.pptx",
     "단계 표기(착수/제안/완료) 없음, 파트명만 기재됨",
     ""),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\전동화&차량개발교육파트\(그룹사 현대케피코) 26년_RnD 직무교육 기획 및 운영_전차_착수보고.pptx",
     "단계 표기가 '_착수보고'로 표준(_착수)과 다름",
     "같은 폴더의 '(그룹사 현대제철) 26년_RnD 직무교육 기획 및 운영_전차_착수.pptx'와 프로젝트명이 완전히 동일 — 회사명만 다름. 복사 후 회사명만 안 바꾼 오기 가능성 있어 확인 필요"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\전동화&차량개발교육파트\(현대트랜시스) 2026년_직무 교육 과정 운영_전차_착수보고_최종.pptx",
     "연도 표기가 '2026년'으로 표준(26년)과 다름, 단계 표기가 '_착수보고_최종'으로 표준(_착수)과 다름",
     "같은 폴더에 정상본 '(현대트랜시스) 26년_직무교육운영_전차_착수.pptx' 존재 — 사실상 중복 파일로 추정, 담당자 확인 필요"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\전동화&차량개발교육파트\(협력사 셰플러 코리아) 26년_전동화 교육 운영_전차_착수보고.pptx",
     "단계 표기가 '_착수보고'로 표준(_착수)과 다름",
     ""),
]

r = 3
for row in rows:
    for i, v in enumerate(row, start=1):
        cell = ws.cell(r, i, v)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        cell.fill = FILL
    r += 1

wb.save(path)
print("시트 추가 완료:", SHEET_NAME)
print("전체 시트:", wb.sheetnames)
