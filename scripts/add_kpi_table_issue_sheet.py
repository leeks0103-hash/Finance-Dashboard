import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_NOTABLE = PatternFill("solid", start_color="E0E0E0", end_color="E0E0E0")   # 테이블 없음
FILL_OLDFMT  = PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0")   # 구버전 템플릿(데이터 오매핑)
FILL_OK      = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")   # 정상 확인

path = r"C:\Users\aaa\coding\dashboard\data\KPI_코드중복_확인요청.xlsx"
wb = load_workbook(path)

SHEET_NAME = "KPI 테이블 없음·양식다름"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)

headers = ["파일명", "문제 유형", "상세"]
widths = [55, 22, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1, "KPI 슬라이드 테이블 없음 / 양식 다름 확인")
c.font = TITLE_FONT
ws.row_dimensions[r].height = 22
r += 2

for i, h in enumerate(headers, start=1):
    cell = ws.cell(r, i, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
r += 1

rows = [
    ("[그룹사 현대위아] 26년_HDAT 사전,사후 학습 콘텐츠 제공_PM_[착수].pptx",
     "KPI 테이블 없음",
     "'KPI/경영현황' 제목이 붙은 슬라이드 자체가 PPT에 없음", FILL_NOTABLE),
    ("[그룹사 현대위아] 26년_HDAT 사전,사후 학습 콘텐츠 제공_PM_[제안].pptx",
     "KPI 테이블 없음",
     "'KPI/경영현황' 제목이 붙은 슬라이드 자체가 PPT에 없음", FILL_NOTABLE),
    ("[대학 세종대 RISE 사업단] 26년_부트캠프_신사업_[제안].pptx",
     "KPI 테이블 없음",
     "'KPI/경영현황' 제목이 붙은 슬라이드 자체가 PPT에 없음", FILL_NOTABLE),
    ("[대학 세종대 RISE 사업단] 26년_부트캠프_신사업_[착수].pptx",
     "KPI 테이블 없음",
     "'KPI/경영현황' 제목이 붙은 슬라이드 자체가 PPT에 없음", FILL_NOTABLE),
    ("[대학 MEGAversity] 26년_미래자동차혁신부품분야 교육과정 제공_신사업_[착수].pptx",
     "KPI 테이블 없음 (AIP 암호화)",
     "AIP 암호화 파일 — win32com으로 복호화 후 재확인해도 'COM 추출 총 0건' — 실제로 매칭되는 표가 없음", FILL_NOTABLE),
    ("[정부 교육부] 26년_매치업_X-AI 교육과정 개발ㆍ운영_PM_[제안].pptx",
     "구버전 템플릿 (데이터 오매핑)",
     "표준 8개 지표가 아니라 10개 지표(학습인원 타본부/그룹사/협력사/기타 포함)의 구버전 템플릿. "
     "추출 스크립트가 행 번호로만 읽어서 3번째 행부터 전부 엉뚱한 지표 컬럼에 값이 들어감. "
     "실제로 'AI교육_적절성_사업계획'에 14157(실제로는 '현대차R&D분야 外 교육 운영실적(교육인원)' 값)이 들어가 있음 — "
     "예전에 'PPT 원본 오입력'으로 알려졌던 그 14157 이상값의 진짜 원인.", FILL_OLDFMT),
    ("[정부 경기TP] 26년_친환경차 부품개발 인력양성 사업_신사업_[제안].pptx",
     "구버전 템플릿 (데이터 오매핑)",
     "위와 동일한 구버전 템플릿(10개 지표) — 같은 방식으로 데이터 오매핑 발생", FILL_OLDFMT),
    ("(그룹사 기아 Autoland) 26년_일반직 S-OJT 매뉴얼 제작 프로젝트_SW_중간.pptx",
     "정상",
     "표준 8개 지표 템플릿 그대로, 구조 이상 없음", FILL_OK),
]

for name, issue, detail, fill in rows:
    cell1 = ws.cell(r, 1, name)
    cell2 = ws.cell(r, 2, issue)
    cell3 = ws.cell(r, 3, detail)
    for cell in (cell1, cell2, cell3):
        cell.font = BODY_FONT
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    lines = max(1, (len(detail) // 45) + detail.count("\n") + 1)
    ws.row_dimensions[r].height = max(30, 15 * lines)
    r += 1

wb.save(path)
print("시트 추가 완료:", SHEET_NAME)
print("전체 시트:", wb.sheetnames)
