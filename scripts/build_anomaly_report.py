import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CAT_FILL = {
    "collision": PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0"),
    "old_tpl": PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
    "aip_fail": PatternFill("solid", start_color="E0E0E0", end_color="E0E0E0"),
    "note_header": PatternFill("solid", start_color="DDEEFF", end_color="DDEEFF"),
}

wb = Workbook()

def style_sheet(ws, headers, col_widths, title):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(1, 1, title)
    c.font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(2, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

def add_row(ws, row_idx, values, fill=None):
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row_idx, i, v)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        if fill:
            cell.fill = fill

# ---------------- Sheet 1: 종합 요약 ----------------
ws1 = wb.active
ws1.title = "종합요약"
headers1 = ["구분", "건수", "설명", "조치"]
style_sheet(ws1, headers1, [22, 8, 55, 45], "재무 PPT 추출 이상 항목 종합 요약 (2026-08-10)")
summary_rows = [
    ("코드 중복 충돌 (진짜 다른 프로젝트)", 1, "경신전선 착수(E105600126030001)가 HDAT-DA Fast Track과 코드 중복 → 경신전선 데이터 소실", "제외 결정됨 (담당자 코드 재발급 필요, 스크립트 로직 변경 안 함)"),
    ("비고 필터링 오탐 (신규 발견)", 1, "H모빌리티클래스 착수/중간 — 정상 비고 텍스트에 '경상 이익' 문구가 포함되어 필터가 통째로 삭제함", "스크립트 clean_note_value() 필터 로직 수정 필요"),
    ("구 템플릿 (제목 불일치, 데이터 미반영)", 5, "슬라이드 제목이 '수익성 분석' 등으로 되어 있어 표는 있지만 스크립트가 인식 못함", "담당자가 최신 템플릿(제목에 정확한 키워드 포함)으로 재제출 필요"),
    ("AIP 암호화 - 진짜 복호화 실패", 2, "win32com 자동화로도 복호화 실패 (파일 손상 또는 강한 보호)", "IT/파일 소유자에게 AIP 보호 해제 요청 필요"),
    ("비고 헤더 텍스트 상이", 23, "표 마지막 컬럼 헤더가 '비고'가 아니라 '차이 설명'/'차이설명'/'세부 내용'/'이익률' 등으로 되어있음 (추출 자체는 정상 동작, fallback으로 처리됨)", "표준 템플릿 컬럼명 '비고'로 통일 요청 (필수는 아님, 정리 차원)"),
]
r = 3
for row in summary_rows:
    add_row(ws1, r, row)
    r += 1
ws1.row_dimensions[2].height = 30

# ---------------- Sheet 2: 코드충돌 및 비고유실 ----------------
ws2 = wb.create_sheet("01_코드충돌_비고유실")
headers2 = ["코드", "구분", "채택된 파일 (현재 남은 데이터)", "충돌/유실된 파일", "문제 유형", "상세"]
style_sheet(ws2, headers2, [20, 10, 42, 42, 20, 45], "코드 중복 충돌 및 비고 유실 사례")
rows2 = [
    ("E105600126030001", "착수", "(HMC 타본부) 26년_인재개발원 HDAT-DA Fast Track_AI_완료.pptx / _착수.pptx",
     "(협력사 경신전선) 26년_AI 교육_AI_착수.pptx", "코드 중복(진짜 다른 프로젝트)",
     "경신전선 담당자가 HDAT-DA 프로젝트 코드를 잘못 기재 → 경신전선 재무데이터 전체 소실. 코드 재발급 필요 (제외 결정됨)"),
    ("E053600125120001", "착수", "(HMC RnD) 26년_H모빌리티클래스_미모_중간.pptx (자체 파일 내)", "동일 파일 내 자체 문제",
     "비고 필터링 오탐(버그)", "원본 비고='- e콘텐츠 수강료 (0.8억) 는 경상 이익으로 반영' → '경상 이익' 문구가 필터 패턴과 겹쳐 공란으로 삭제됨"),
    ("E053600125120001", "중간", "(HMC RnD) 26년_H모빌리티클래스_미모_중간.pptx (자체 파일 내)", "동일 파일 내 자체 문제",
     "비고 필터링 오탐(버그)", "위와 동일한 원인으로 비고 소실"),
]
r = 3
for row in rows2:
    add_row(ws2, r, row, CAT_FILL["collision"])
    r += 1

# ---------------- Sheet 3: 구 템플릿 (데이터 미반영) ----------------
ws3 = wb.create_sheet("02_구템플릿_미반영")
headers3 = ["파일 위치", "실제 슬라이드 제목", "비고"]
style_sheet(ws3, headers3, [70, 30, 45], "구 템플릿 사용 — 표는 있으나 제목 키워드 불일치로 미반영")
rows3 = [
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\교육사업PM파트\[협력사 한국타이어앤테크놀로지] 26년_기술 콘텐츠 제공_PM_[착수]_수정.pptx", "수익성 분석", "코드=H043600126030001, 착수, 매출 10,500,000원 데이터 확인됨"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\신사업기획파트\[대학 세종대 RISE 사업단] 26년_부트캠프_신사업_[착수].pptx", "수익성 분석", "완전히 다른 컬럼 구조(PMS 프로젝트 코드 등 20컬럼) — 자동 변환 불가, 수기 확인 필요"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\신사업기획파트\[대학 세종대 RISE 사업단] 26년_부트캠프_신사업_[제안].pptx", "(수익성 분석 계열, 별도 표)", "구분/총매출/직접원가/인건비/제경비/영업이익 구조 — 표준 컬럼과 다름"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\교육사업PM파트\[그룹사 현대위아] 26년_HDAT 사전,사후 학습 콘텐츠 제공_PM_[제안].pptx", "수익성 분석", "완전히 다른 컬럼 구조(PMS 프로젝트 코드 등 20컬럼) — 자동 변환 불가"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\교육사업PM파트\[그룹사 현대위아] 26년_HDAT 사전,사후 학습 콘텐츠 제공_PM_[착수].pptx", "수익성 분석", "완전히 다른 컬럼 구조(PMS 프로젝트 코드 등 20컬럼) — 자동 변환 불가"),
]
r = 3
for row in rows3:
    add_row(ws3, r, row, CAT_FILL["old_tpl"])
    r += 1

# ---------------- Sheet 4: AIP 암호화 실패 ----------------
ws4 = wb.create_sheet("03_AIP_암호화실패")
headers4 = ["파일 위치", "상태"]
style_sheet(ws4, headers4, [90, 30], "AIP 암호화 — win32com 자동화로도 복호화 실패 (진짜 실패)")
rows4 = [
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\미래모빌리티교육파트\(HMC RnD) 26년_수소특화교육_미모_제안.pptx", "복호화 실패 (SaveAs/슬라이드 복붙 방식 모두 실패)"),
    (r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집\미래모빌리티교육파트\(HMC 국내생산담당) 역량인증제 직무체계 검증 및 문항개발 워크샵_미모_착수.pptx", "복호화 실패 (SaveAs/슬라이드 복붙 방식 모두 실패)"),
]
r = 3
for row in rows4:
    add_row(ws4, r, row, CAT_FILL["aip_fail"])
    r += 1

# ---------------- Sheet 5: 비고 헤더 텍스트 상이 ----------------
ws5 = wb.create_sheet("04_비고헤더텍스트상이")
headers5 = ["파일명", "실제 마지막 컬럼 헤더"]
style_sheet(ws5, headers5, [75, 25], "표 마지막 컬럼 헤더가 '비고'가 아닌 파일 (추출 자체는 정상 동작)")
variant_rows = [
    ("(HMC RnD 통합) 26년_직무 특화 역량 향상 교육 운영_통합관리_착수.pptx", "차이 설명"),
    ("(그룹사 현대오토에버) 26년_오토에버 사내 과정 운영_SW_착수.pptx", "차이 설명"),
    ("[정부 교육부] 26년_매치업_X-AI 교육과정 개발ㆍ운영_PM_[제안].pptx", "차이 설명"),
    ("(HMC 국내생산담당) 역량인증제 직무체계 검증 및 문항개발 워크샵_미모_제안.pptx", "차이 설명"),
    ("(HMC 제조솔루션본부) 26년 제조SW아카데미_미모_중간.pptx", "차이 설명"),
    ("(HMC 제조솔루션본부) 26년_부트-캠퍼스_미모_제안.pptx", "차이 설명"),
    ("(HMC 제조솔루션본부) 26년_부트-캠퍼스_미모_착수.pptx", "차이 설명"),
    ("(HMC 제조솔루션본부) 26년_역량 진단 평가 개발 및 운영_미모_제안.pptx", "차이 설명"),
    ("(HMC 제조솔루션본부) 26년_제조SW아카데미_미모_제안.pptx", "차이 설명"),
    ("(HMC 제조솔루션본부) 26년_제조SW아카데미_미모_착수.pptx", "차이 설명"),
    ("[경기테크노파크] 친환경차 부품개발 인력양성 위탁교육 착수보고_260513.pptx", "차이 설명"),
    ("[정부 경기TP] 26년_친환경차 부품개발 인력양성 사업_신사업_[제안].pptx", "차이 설명"),
    ("[정부 경기TP] 26년_친환경차 부품개발 인력양성 위탁교육_신사업_[착수].pptx", "차이 설명"),
    ("[정부 서울시] 25년_새싹2기 차세대 모빌리티 임베디드 AI 엔지니어 양성_신사업_[완료].pptx", "이익률 (비고 컬럼 자체 없음 - 별도 확인 필요)"),
    ("[정부 서울시] 26년_새싹3기 차세대 모빌리티 임베디드 AI 엔지니어 양성_신사업_[제안].pptx", "이익률 (비고 컬럼 자체 없음 - 별도 확인 필요)"),
    ("[정부 평택산업진흥원] 26년_제조 AI 인재양성 사업_신사업_[제안].pptx", "세부 내용"),
    ("[정부 한국AI로봇산업협회] 26년_산업전문인력 AI 역량강화 프로젝트 (로봇산업 리더교육)_신사업_[착수].pptx", "차이설명"),
    ("[정부 한국전자정보통신산업진흥회] 26년_산업전문인력 AI역량강화 프로젝트_신사업_[사전검토].pptx", "차이 설명"),
    ("[정부 한국전자정보통신산업진흥회] 26년_산업전문인력 AI역량강화 프로젝트_신사업_[제안].pptx", "차이설명"),
    ("[정부 한국전자정보통신산업진흥회] 26년_산업전문인력 AI역량강화 프로젝트_신사업_[제안]_미선정.pptx", "차이설명"),
    ("[현대자동차] 26년_바디시험 기술역량 강화(심화코칭) 프로그램_신사업_[제안].pptx", "이익률 (비고 컬럼 자체 없음 - 별도 확인 필요)"),
    ("[현대자동차] 26년_바디시험 기술역량 강화(심화코칭) 프로그램_신사업_[착수].pptx", "이익률 (비고 컬럼 자체 없음 - 별도 확인 필요)"),
    ("경기TP 친환경차 부품개발 인력양성 사업 제안보고_260410_최종.pptx", "차이 설명"),
]
r = 3
for row in variant_rows:
    add_row(ws5, r, row, CAT_FILL["note_header"])
    r += 1

out_path = r"C:\Users\aaa\coding\dashboard\data\재무PPT_이상항목_종합보고서.xlsx"
wb.save(out_path)
print("저장 완료:", out_path)
