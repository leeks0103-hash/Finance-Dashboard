import os
import re
import sys
import time
import traceback
from datetime import datetime

import pythoncom
import win32com.client
from openpyxl import Workbook, load_workbook

# =========================
# 설정값
# =========================
# 환경변수 EXTRACT_BASE_DIR 우선 사용 — compare_and_update.py가 자동 주입
BASE_DIR = os.environ.get(
    "EXTRACT_BASE_DIR",
    r"C:\Users\aaa\Desktop\기술교육실_프로젝트 보고서 수집",
)

# 출력 엑셀: 프로젝트 data/ 폴더로 저장
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR   = os.path.join(os.path.dirname(_SCRIPT_DIR), "data")
os.makedirs(_DATA_DIR, exist_ok=True)

TARGET_EXCEL = os.path.join(_DATA_DIR, "재무관점 필수 데이터 추출.xlsx")
TARGET_SHEET = "취합"
HISTORY_SHEET = "처리이력"
TITLE_KEYWORD = "[내부용①] 재무관점 필수 데이터"
SUPPORTED_EXTENSIONS = {".ppt", ".pptx"}
LOG_FILE = os.path.join(_DATA_DIR, "extract_financial_ppt.log")

EXCLUDE_FILENAMES = {
    "테스트 입니다.pptx",
}

# 기존 저장 데이터 삭제 후 전체 재추출
FORCE_REPROCESS = True
RESET_OUTPUT_ON_START = True

PART_KEYWORDS = ["신사업", "PM", "전차", "미모", "AI", "SW", "K뉴딜TF"]

OUTPUT_HEADERS = [
    "프로젝트 코드",     # 1
    "연도",             # 2
    "파트명",           # 3
    "구분",             # 4
    "총매출",           # 5
    "지출합계",         # 6
    "직접원가",         # 7
    "직접 인건비",      # 8
    "공통원가/관리비",  # 9
    "경상 이익",        # 10
    "이익율",           # 11 (float, 표시형식 0.00)
    "비고",             # 12
    "원본파일명",       # 13
    "원본수정일시",     # 14
    "반영일시",         # 15
]

HISTORY_HEADERS = [
    "파일식별키",
    "파일명",
    "전체경로",
    "최종수정일시",
    "파일크기",
    "처리일시",
    "처리상태",
    "메시지",
]

# =========================
# 로그
# =========================
def log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


# =========================
# 공통 유틸
# =========================
def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def file_mtime_str(path):
    return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")


def build_file_signature(path):
    name = os.path.basename(path)
    mtime = str(int(os.path.getmtime(path)))
    size = str(os.path.getsize(path))
    return f"{name}|{mtime}|{size}"


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_cell_value(value):
    text = normalize_text(value)
    text = text.replace("\r", "").replace("\x0b", "\n").replace("\n", " ").strip()
    if text in ("", "-"):
        return "0"
    return text


def normalize_numeric_text(value):
    """
    숫자 변환 전 공통 정리
    - 쉼표/공백 제거
    - 퍼센트 제거: 5.0% -> 5.0
    - 괄호 음수 처리
    - 문자열 속 첫 숫자 패턴 추출
    """
    text = clean_cell_value(value)

    if text in ("", "-", "0"):
        return "0"

    text = str(text).strip()
    text = text.replace(",", "")
    text = text.replace(" ", "")
    text = text.replace("％", "%")
    text = text.replace("（", "(").replace("）", ")")

    text = text.replace("▲", "-")
    text = text.replace("△", "-")
    text = text.replace("▼", "-")
    text = text.replace("▽", "-")
    text = text.replace("＋", "+")
    text = text.replace("－", "-")

    text = re.sub(r"(?i)%p$", "", text)
    text = re.sub(r"(?i)%$", "", text)
    text = re.sub(r"(?i)percent$", "", text)
    text = re.sub(r"(?i)퍼센트$", "", text)

    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if match:
        return match.group(0)

    return "0"


def to_number(value):
    text = normalize_numeric_text(value)
    try:
        num = float(text)
        if num.is_integer():
            return int(num)
        return num
    except Exception:
        return 0


def to_float(value):
    text = normalize_numeric_text(value)
    try:
        return float(text)
    except Exception:
        log(f"[11열 이익율 변환실패] 원본값={repr(value)} / 정규화값={repr(text)}")
        return 0.0


def safe_str(value):
    return str(clean_cell_value(value))


def is_temp_or_hidden_office_file(filename):
    return filename.startswith("~$")


def is_excluded_file(filename):
    return filename.strip().lower() in {f.lower() for f in EXCLUDE_FILENAMES}


def extract_year_from_filename(filename):
    """
    예:
    [현대자동차] 26년_열에너지통합개발 역량 강화 프로젝트_신사업_[착수].pptx
    -> 2026
    """
    name = os.path.splitext(os.path.basename(filename))[0]

    match = re.search(r"(?<!\d)(\d{2})년(?!\d)", name)
    if match:
        yy = int(match.group(1))
        return 2000 + yy

    match = re.search(r"(?<!\d)(20\d{2})(?!\d)", name)
    if match:
        return int(match.group(1))

    return "-"


def extract_part_from_filename(filename):
    """
    파일명에 PART_KEYWORDS 중 포함되는 첫 번째 키워드 반환
    없으면 '-'
    """
    name = os.path.splitext(os.path.basename(filename))[0]

    for keyword in PART_KEYWORDS:
        if keyword in name:
            return keyword

    return "-"


# =========================
# 엑셀 초기화 / 로드
# =========================
def ensure_workbook(excel_path):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

    if TARGET_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TARGET_SHEET)
        ws.append(OUTPUT_HEADERS)

    if HISTORY_SHEET not in wb.sheetnames:
        hs = wb.create_sheet(HISTORY_SHEET)
        hs.append(HISTORY_HEADERS)

    apply_sheet_layout(wb[TARGET_SHEET])
    wb.save(excel_path)
    return wb


def reset_output_sheets(wb):
    """
    기존 저장 데이터 삭제 후 새 규칙으로 전체 재추출
    - 취합 시트: 헤더만 남기고 삭제
    - 처리이력 시트: 헤더만 남기고 삭제
    """
    data_ws = get_or_create_sheet(wb, TARGET_SHEET, OUTPUT_HEADERS)
    history_ws = get_or_create_sheet(wb, HISTORY_SHEET, HISTORY_HEADERS)

    if data_ws.max_row > 1:
        data_ws.delete_rows(2, data_ws.max_row - 1)

    if history_ws.max_row > 1:
        history_ws.delete_rows(2, history_ws.max_row - 1)

    for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
        data_ws.cell(1, col_idx, header)

    for col_idx, header in enumerate(HISTORY_HEADERS, start=1):
        history_ws.cell(1, col_idx, header)

    apply_sheet_layout(data_ws)
    wb.save(TARGET_EXCEL)


def get_or_create_sheet(wb, sheet_name, headers):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

    ws = wb[sheet_name]

    if ws.max_row == 1 and all(ws.cell(1, c).value is None for c in range(1, len(headers) + 1)):
        for i, h in enumerate(headers, start=1):
            ws.cell(1, i, h)

    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.delete_rows(1, 1)
        ws.append(headers)

    return ws


def apply_sheet_layout(data_ws):
    widths = {
        1: 18,   # 프로젝트 코드
        2: 10,   # 연도
        3: 14,   # 파트명
        4: 14,   # 구분
        5: 14,   # 총매출
        6: 14,   # 지출합계
        7: 14,   # 직접원가
        8: 16,   # 직접 인건비
        9: 20,   # 공통원가/관리비
        10: 14,  # 경상 이익
        11: 12,  # 이익율
        12: 24,  # 비고
        13: 28,  # 원본파일명
        14: 20,  # 원본수정일시
        15: 20,  # 반영일시
    }

    for col_idx, width in widths.items():
        col_letter = data_ws.cell(row=1, column=col_idx).column_letter
        data_ws.column_dimensions[col_letter].width = width

    data_ws.freeze_panes = "A2"


def apply_number_formats(data_ws):
    for row_num in range(2, data_ws.max_row + 1):
        data_ws.cell(row=row_num, column=11).number_format = "0.00"


# =========================
# 처리이력 관리
# =========================
def load_processed_signatures(history_ws):
    processed = set()
    for row in history_ws.iter_rows(min_row=2, values_only=True):
        signature = row[0]
        status = row[6] if len(row) >= 7 else None
        if signature and status == "SUCCESS":
            processed.add(signature)
    return processed


def append_history(history_ws, signature, path, status, message=""):
    history_ws.append([
        signature,
        os.path.basename(path),
        path,
        file_mtime_str(path) if os.path.exists(path) else "",
        os.path.getsize(path) if os.path.exists(path) else "",
        now_str(),
        status,
        message,
    ])


# =========================
# 취합 시트 인덱스 관리
# =========================
def build_data_index(data_ws):
    index_map = {}
    for row_num in range(2, data_ws.max_row + 1):
        key_project = normalize_text(data_ws.cell(row=row_num, column=1).value)
        key_year = normalize_text(data_ws.cell(row=row_num, column=2).value)
        key_part = normalize_text(data_ws.cell(row=row_num, column=3).value)
        key_gubun = normalize_text(data_ws.cell(row=row_num, column=4).value)
        if key_project or key_gubun:
            index_map[(key_project, key_year, key_part, key_gubun)] = row_num
    return index_map


def upsert_rows(data_ws, rows):
    index_map = build_data_index(data_ws)
    inserted = 0
    updated = 0

    for row_data in rows:
        key = (
            normalize_text(row_data[0]),
            normalize_text(row_data[1]),
            normalize_text(row_data[2]),
            normalize_text(row_data[3]),
        )

        if key in index_map:
            row_num = index_map[key]
            for col_idx, value in enumerate(row_data, start=1):
                data_ws.cell(row=row_num, column=col_idx, value=value)
            data_ws.cell(row=row_num, column=11).number_format = "0.00"
            updated += 1
        else:
            data_ws.append(row_data)
            new_row_num = data_ws.max_row
            data_ws.cell(row=new_row_num, column=11).number_format = "0.00"
            index_map[key] = new_row_num
            inserted += 1

    return inserted, updated


# =========================
# PowerPoint 읽기
# =========================
def get_slide_title_text(slide):
    try:
        if slide.Shapes.HasTitle:
            title_shape = slide.Shapes.Title
            if title_shape is not None and hasattr(title_shape, "TextFrame"):
                return normalize_text(title_shape.TextFrame.TextRange.Text)
    except Exception:
        pass

    texts = []
    try:
        for i in range(1, slide.Shapes.Count + 1):
            shp = slide.Shapes(i)
            try:
                if shp.HasTextFrame and shp.TextFrame.HasText:
                    txt = normalize_text(shp.TextFrame.TextRange.Text)
                    if txt:
                        texts.append(txt)
            except Exception:
                continue
    except Exception:
        pass

    return " ".join(texts)


def extract_first_table_from_slide(slide):
    for i in range(1, slide.Shapes.Count + 1):
        shp = slide.Shapes(i)
        try:
            if shp.HasTable:
                return shp.Table
        except Exception:
            continue
    return None


def extract_rows_from_table(table, source_file, source_mtime):
    extracted = []

    row_count = table.Rows.Count
    col_count = table.Columns.Count

    if col_count < 10:
        raise ValueError(f"표 컬럼 수가 10개 미만입니다. 현재 컬럼 수: {col_count}")

    # 재경비 컬럼 감지: 11열 이상이면 col 8 위치에 재경비 존재로 판단
    # 구 템플릿(10열): 프로젝트코드|구분|총매출|지출합계|직접원가|인건비|공통원가|경상이익|이익율|비고
    # 신 템플릿(11열): 프로젝트코드|구분|총매출|지출합계|직접원가|인건비|공통원가|재경비|경상이익|이익율|비고
    has_jaegyungbi = col_count >= 11
    read_cols = min(col_count, 13)  # 최대 13열까지 읽기

    source_filename = os.path.basename(source_file)
    extracted_year = extract_year_from_filename(source_filename)
    extracted_part = extract_part_from_filename(source_filename)

    for r in range(2, row_count + 1):
        raw_values = []
        for c in range(1, read_cols + 1):
            cell_text = ""
            try:
                cell = table.Cell(r, c)
                cell_text = cell.Shape.TextFrame.TextRange.Text
            except Exception:
                cell_text = ""
            raw_values.append(cell_text)

        # 재경비 컬럼이 있으면 index 7(0-based)을 건너뛰어 뒤 컬럼 인덱스 조정
        offset = 1 if has_jaegyungbi else 0

        row_data = []

        row_data.append(safe_str(raw_values[0]))              # 1열: 프로젝트 코드
        row_data.append(extracted_year)                        # 2열: 연도
        row_data.append(extracted_part)                        # 3열: 파트명
        row_data.append(safe_str(raw_values[1]))              # 4열: 구분
        row_data.append(to_number(raw_values[2]))             # 5열: 총매출
        row_data.append(to_number(raw_values[3]))             # 6열: 지출합계
        row_data.append(to_number(raw_values[4]))             # 7열: 직접원가
        row_data.append(to_number(raw_values[5]))             # 8열: 직접 인건비
        row_data.append(to_number(raw_values[6]))             # 9열: 공통원가/관리비
        # raw_values[7] = 재경비 (신 템플릿) → skip
        row_data.append(to_number(raw_values[7 + offset]))    # 10열: 경상 이익

        raw_col11 = raw_values[8 + offset]                    # 11열: 이익율
        converted_col11 = to_float(raw_col11)
        log(
            f"[11열 이익율 확인] 파일={source_filename} / 행={r} / "
            f"연도={extracted_year} / 파트명={extracted_part} / "
            f"재경비컬럼={'있음' if has_jaegyungbi else '없음'} / "
            f"원본={repr(raw_col11)} / 변환={converted_col11}"
        )
        row_data.append(converted_col11)                      # 11열: 이익율

        row_data.append(safe_str(raw_values[9 + offset]))    # 12열: 비고

        key_project = normalize_text(row_data[0])
        key_gubun = normalize_text(row_data[3])
        if key_project in ("", "0") and key_gubun in ("", "0"):
            continue

        row_data.append(source_filename)              # 13열: 원본파일명
        row_data.append(source_mtime)                 # 14열: 원본수정일시
        row_data.append(now_str())                    # 15열: 반영일시

        extracted.append(row_data)

    return extracted


def open_presentation_with_retry(ppt_app, ppt_path, max_retries=2, wait_seconds=1):
    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            return ppt_app.Presentations.Open(ppt_path, True, False, False)
        except Exception as e:
            last_error = e
            if attempt < max_retries:
                log(f"프레젠테이션 열기 재시도 {attempt}/{max_retries - 1}: {ppt_path}")
                time.sleep(wait_seconds)
            else:
                raise last_error


def extract_financial_rows_from_ppt(ppt_app, ppt_path):
    rows_all = []
    presentation = None

    try:
        presentation = open_presentation_with_retry(ppt_app, ppt_path)
        src_mtime = file_mtime_str(ppt_path)

        for slide_idx in range(1, presentation.Slides.Count + 1):
            slide = presentation.Slides(slide_idx)
            title_text = get_slide_title_text(slide)

            if TITLE_KEYWORD not in title_text:
                continue

            table = extract_first_table_from_slide(slide)
            if table is None:
                log(f"표 없음 - 파일: {ppt_path}, 슬라이드: {slide_idx} / 키워드: {TITLE_KEYWORD}")
                continue

            slide_rows = extract_rows_from_table(table, ppt_path, src_mtime)
            rows_all.extend(slide_rows)

        return rows_all

    finally:
        if presentation is not None:
            try:
                presentation.Close()
            except Exception:
                pass


# =========================
# 파일 탐색
# =========================
def find_target_ppt_files(base_dir, processed_signatures):
    candidates = []
    excluded_count = 0

    for root, _, files in os.walk(base_dir):
        for name in files:
            if is_temp_or_hidden_office_file(name):
                continue

            ext = os.path.splitext(name)[1].lower()
            if ext not in SUPPORTED_EXTENSIONS:
                continue

            full_path = os.path.join(root, name)

            if os.path.abspath(full_path).lower() == os.path.abspath(TARGET_EXCEL).lower():
                continue

            if is_excluded_file(name):
                log(f"[제외] 제외 파일명 목록에 포함되어 건너뜀: {full_path}")
                excluded_count += 1
                continue

            signature = build_file_signature(full_path)

            if not FORCE_REPROCESS and signature in processed_signatures:
                continue

            candidates.append(full_path)

    candidates.sort(key=lambda x: os.path.getmtime(x), reverse=True)

    if excluded_count > 0:
        log(f"[제외 요약] 제외된 파일 수: {excluded_count}개")

    return candidates


# =========================
# PowerPoint 앱
# =========================
def create_powerpoint_app():
    app = win32com.client.DispatchEx("PowerPoint.Application")
    app.Visible = 1
    time.sleep(1)
    return app


# =========================
# 메인 처리
# =========================
def main():
    if not os.path.isdir(BASE_DIR):
        log(f"[오류] 대상 폴더를 찾을 수 없습니다: {BASE_DIR}")
        return 1

    try:
        wb = ensure_workbook(TARGET_EXCEL)
    except Exception as e:
        log(f"[오류] 대상 엑셀 파일을 열거나 생성할 수 없습니다: {e}")
        return 1

    if RESET_OUTPUT_ON_START:
        log("[초기화] 기존 저장 데이터 삭제 후 전체 재추출을 시작합니다.")
        reset_output_sheets(wb)

    data_ws = get_or_create_sheet(wb, TARGET_SHEET, OUTPUT_HEADERS)
    history_ws = get_or_create_sheet(wb, HISTORY_SHEET, HISTORY_HEADERS)
    apply_sheet_layout(data_ws)

    processed_signatures = load_processed_signatures(history_ws)
    target_files = find_target_ppt_files(BASE_DIR, processed_signatures)

    if not target_files:
        apply_number_formats(data_ws)
        wb.save(TARGET_EXCEL)
        log("[완료] 처리할 신규/변경 PowerPoint 파일이 없습니다.")
        return 0

    pythoncom.CoInitialize()
    ppt_app = None

    total_files = 0
    total_success_files = 0
    total_fail_files = 0
    total_inserted = 0
    total_updated = 0
    total_extracted = 0

    try:
        ppt_app = create_powerpoint_app()
        log(f"[시작] 처리 대상 파일 수: {len(target_files)}")
        log(f"[시작] 추출 키워드: {TITLE_KEYWORD}")
        log(f"[시작] 제외 파일명 목록: {', '.join(EXCLUDE_FILENAMES) if EXCLUDE_FILENAMES else '없음'}")
        log(f"[시작] 파트명 키워드 목록: {', '.join(PART_KEYWORDS)}")
        log(f"[시작] 출력 파일: {TARGET_EXCEL}")

        for ppt_path in target_files:
            signature = build_file_signature(ppt_path)
            total_files += 1

            try:
                log(f"[처리중] {ppt_path}")
                extracted_rows = extract_financial_rows_from_ppt(ppt_app, ppt_path)

                if not extracted_rows:
                    append_history(
                        history_ws,
                        signature,
                        ppt_path,
                        "SUCCESS",
                        f"조건에 맞는 '{TITLE_KEYWORD}' 표 데이터 없음"
                    )
                    total_success_files += 1
                    log("  └ 추출 데이터 없음")
                    continue

                inserted, updated = upsert_rows(data_ws, extracted_rows)
                total_inserted += inserted
                total_updated += updated
                total_extracted += len(extracted_rows)
                total_success_files += 1

                append_history(
                    history_ws,
                    signature,
                    ppt_path,
                    "SUCCESS",
                    f"추출 {len(extracted_rows)}건 / 추가 {inserted}건 / 갱신 {updated}건"
                )

                log(f"  └ 추출 {len(extracted_rows)}건 / 추가 {inserted}건 / 갱신 {updated}건")

            except Exception as e:
                total_fail_files += 1
                err_msg = f"{type(e).__name__}: {e}"
                append_history(history_ws, signature, ppt_path, "FAIL", err_msg)
                log(f"  └ 오류: {err_msg}")
                log(traceback.format_exc())

        apply_number_formats(data_ws)
        apply_sheet_layout(data_ws)
        wb.save(TARGET_EXCEL)

        log("[완료]")
        log(f"- 전체 파일 수: {total_files}")
        log(f"- 성공 파일 수: {total_success_files}")
        log(f"- 실패 파일 수: {total_fail_files}")
        log(f"- 총 추출 건수: {total_extracted}")
        log(f"- 추가 건수: {total_inserted}")
        log(f"- 갱신 건수: {total_updated}")
        log(f"- 결과 파일: {TARGET_EXCEL}")
        log(f"- 로그 파일: {LOG_FILE}")

        return 0

    finally:
        try:
            apply_number_formats(data_ws)
            wb.save(TARGET_EXCEL)
        except Exception:
            pass

        if ppt_app is not None:
            try:
                ppt_app.Quit()
            except Exception:
                pass

        pythoncom.CoUninitialize()


if __name__ == "__main__":
    sys.exit(main())
