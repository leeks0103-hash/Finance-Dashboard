import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_DUP = PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0")
FILL_CHECK = PatternFill("solid", start_color="DDEEFF", end_color="DDEEFF")

wb = Workbook()
ws = wb.active
ws.title = "KPI 코드 중복"

widths = [20, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
c = ws.cell(r, 1, "KPI 슬라이드 코드 중복 — 서로 다른 프로젝트/확인 필요 건")
c.font = TITLE_FONT
ws.row_dimensions[r].height = 22
r += 2

# ── 섹션 1: 서로 다른 프로젝트끼리 코드 중복 ──
ws.cell(r, 1, "서로 다른 프로젝트끼리 코드 중복 (개별 건)").font = SECTION_FONT
r += 1
for i, h in enumerate(["공유 코드", "겹치는 프로젝트"], start=1):
    cell = ws.cell(r, i, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
r += 1

dup_rows = [
    ("H086600126010001", "상용개발센터 FMEA 연구역량 강화 프로그램 ↔ 신뢰성·강건개발 인증제 운영\n"
     "- (HMC RnD) 26년_상용개발센터 FMEA 연구역량 강화 프로그램_전차_제안.pptx\n"
     "- (HMC RnD) 26년_상용개발센터 FMEA 연구역량 강화 프로그램_전차_착수.pptx\n"
     "- (HMC RnD) 26년_신뢰성·강건개발 인증제 운영_전차_제안.pptx\n"
     "- (HMC RnD) 26년_신뢰성·강건개발 인증제 운영_전차_착수.pptx"),
    ("E017600126040001", "자율주행 특화 ↔ 현대모비스 RnD특화 SW 교육 ↔ R&D 사내교육 (3개 프로젝트)\n"
     "- (그룹사 현대모비스) 26년_자율주행 특화_SW_착수.pptx\n"
     "- (그룹사 현대모비스) 26년_현대모비스 RnD특화 SW 교육_SW_착수.pptx\n"
     "- (그룹사 현대모비스) 26년_R&D 사내교육_SW+전차_착수.pptx"),
    ("H095600126050001", "영남대학교 RISE 사업단 ↔ 새싹3기 차세대 모빌리티 ↔ 외장메커니즘 (3개 프로젝트)\n"
     "- [정부 교육부] 26년_영남대학교 RISE 사업단 교육과정_신사업_[제안]_수정.pptx\n"
     "- [정부 서울시] 26년_새싹3기 차세대 모빌리티 임베디드 AI 엔지니어 양성_신사업_[제안].pptx\n"
     "- [현대자동차] 26년_외장메커니즘 역량 강화 프로젝트_신사업_[착수].pptx"),
]

for code, desc in dup_rows:
    cell1 = ws.cell(r, 1, code)
    cell2 = ws.cell(r, 2, desc)
    for cell in (cell1, cell2):
        cell.font = BODY_FONT
        cell.fill = FILL_DUP
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 15 * (desc.count("\n") + 1) + 6
    r += 1

r += 1

# ── 섹션 2: 확인 필요 (표기 차이일 수 있음) ──
ws.cell(r, 1, "확인 필요 (같은 프로젝트 표기 차이일 수 있음)").font = SECTION_FONT
r += 1
for i, h in enumerate(["공유 코드", "파일"], start=1):
    cell = ws.cell(r, i, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
r += 1

check_desc = ("경기테크노파크/경기TP 친환경차 부품개발 관련 3개 파일 (약칭 차이로 동일 프로젝트일 가능성)\n"
              "- [경기테크노파크] 친환경차 부품개발 인력양성 위탁교육_신사업_착수_260513.pptx\n"
              "- [정부 경기TP] 26년_친환경차 부품개발 인력양성 위탁교육_신사업_[착수].pptx\n"
              "- 경기TP 친환경차 부품개발 인력양성 사업 제안보고_260410_최종.pptx")
cell1 = ws.cell(r, 1, "E117600126020001")
cell2 = ws.cell(r, 2, check_desc)
for cell in (cell1, cell2):
    cell.font = BODY_FONT
    cell.fill = FILL_CHECK
    cell.border = BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.row_dimensions[r].height = 15 * (check_desc.count("\n") + 1) + 6

out_path = r"C:\Users\aaa\coding\dashboard\data\KPI_코드중복_확인요청.xlsx"
wb.save(out_path)
print("저장 완료:", out_path)
