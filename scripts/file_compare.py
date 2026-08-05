import os
import hashlib
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ============================================================
# 📁 경로 설정
# ============================================================
NAS_PATH   = r"X:"       # 기존 파일
LOCAL_PATH = r"D:\24.기술교육사업기획팀\23. 표준 템플릿 데이터 추출 프로젝트\(기술교육실)프로젝트 보고서 수집"   # 새 파일
EXCEL_PATH = r"D:\ngv_dashbord\Finance-Dashboard\data\compare_report.xlsx"  # 결과 저장

# ============================================================
# ⚙️ 옵션 설정
# ============================================================
USE_HASH_CHECK = False

# ✅ 비교 대상 확장자 (ppt, pptx 만 비교)
TARGET_EXTENSIONS = {'.ppt', '.pptx'}

# 무시할 폴더
IGNORE_DIRS = {'$RECYCLE.BIN', 'System Volume Information'}

# ============================================================
# 🎨 엑셀 스타일 정의
# ============================================================
COLOR = {
    "header_bg"   : "1F4E79",
    "header_font" : "FFFFFF",
    "nas_only"    : "D6E4FF",
    "local_only"  : "E2EFDA",
    "modified"    : "FFF2CC",
    "identical"   : "F2F2F2",
    "section_nas" : "2E75B6",
    "section_loc" : "548235",
    "section_mod" : "C55A11",
    "border"      : "BFBFBF",
}

def thin_border():
    side = Side(style='thin', color=COLOR["border"])
    return Border(left=side, right=side, top=side, bottom=side)

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

# ============================================================
# 🔧 유틸리티 함수
# ============================================================
def get_md5(filepath: str, chunk_size: int = 8192) -> str:
    hasher = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while chunk := f.read(chunk_size):
                hasher.update(chunk)
        return hasher.hexdigest()
    except (PermissionError, OSError):
        return "ERROR_CANNOT_READ"

def format_size(size_bytes: int) -> str:
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"

def format_time(timestamp: float) -> str:
    return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')

def is_target_file(path: str) -> bool:
    """✅ ppt / pptx 파일만 비교 대상으로 필터링"""
    return Path(path).suffix.lower() in TARGET_EXTENSIONS

def is_ignored_dir(dirname: str) -> bool:
    return dirname in IGNORE_DIRS

# ============================================================
# 📊 파일 목록 수집
# ============================================================
def collect_files(base_path: str) -> dict:
    """ppt / pptx 파일만 수집"""
    files = {}
    base = Path(base_path)

    if not base.exists():
        print(f"  ⚠️  경로가 존재하지 않습니다: {base_path}")
        return files

    for dirpath, dirnames, filenames in os.walk(base_path):
        # 무시할 디렉토리 제외
        dirnames[:] = [d for d in dirnames if not is_ignored_dir(d)]

        for filename in filenames:
            abs_path = os.path.join(dirpath, filename)
            rel_path = os.path.relpath(abs_path, base_path)

            # ✅ ppt / pptx 확장자만 처리
            if not is_target_file(filename):
                continue

            try:
                stat = os.stat(abs_path)
                files[rel_path] = {
                    'size'   : stat.st_size,
                    'mtime'  : stat.st_mtime,
                    'abspath': abs_path
                }
            except (PermissionError, OSError) as e:
                print(f"  ⚠️  파일 접근 오류: {abs_path} → {e}")
    return files

# ============================================================
# 🔍 폴더 비교
# ============================================================
def compare_folders(nas_path: str, local_path: str) -> dict:
    print("=" * 65)
    print("  📂 폴더 비교 시작")
    print("=" * 65)
    print(f"  🖥️  NAS  경로  : {nas_path}")
    print(f"  💻  로컬 경로  : {local_path}")
    print(f"  🔧  비교 방식  : {'MD5 해시' if USE_HASH_CHECK else '크기 + 수정시간'}")
    print(f"  📎  대상 확장자: {', '.join(sorted(TARGET_EXTENSIONS))}")
    print("=" * 65)

    print("\n  📥 NAS 파일 목록 수집 중...")
    nas_files = collect_files(nas_path)
    print(f"      → {len(nas_files):,}개 파일 발견 (ppt/pptx)")

    print("  📥 로컬 파일 목록 수집 중...")
    local_files = collect_files(local_path)
    print(f"      → {len(local_files):,}개 파일 발견 (ppt/pptx)\n")

    nas_set   = set(nas_files.keys())
    local_set = set(local_files.keys())

    result = {
        'only_in_nas'  : [],
        'only_in_local': [],
        'modified'     : [],
        'identical'    : 0
    }

    # NAS에만 있는 파일
    for rel in sorted(nas_set - local_set):
        info = nas_files[rel]
        result['only_in_nas'].append({
            'path' : rel,
            'size' : format_size(info['size']),
            'mtime': format_time(info['mtime'])
        })

    # 로컬에만 있는 파일
    for rel in sorted(local_set - nas_set):
        info = local_files[rel]
        result['only_in_local'].append({
            'path' : rel,
            'size' : format_size(info['size']),
            'mtime': format_time(info['mtime'])
        })

    # 공통 파일 비교
    common_files = nas_set & local_set
    print(f"  🔄 공통 파일 {len(common_files):,}개 비교 중...")

    for rel in sorted(common_files):
        nas_info   = nas_files[rel]
        local_info = local_files[rel]

        if USE_HASH_CHECK:
            is_different = get_md5(nas_info['abspath']) != get_md5(local_info['abspath'])
            diff_reason  = "MD5 불일치"
        else:
            size_diff    = nas_info['size'] != local_info['size']
            mtime_diff   = abs(nas_info['mtime'] - local_info['mtime']) > 1
            is_different = size_diff or mtime_diff
            reasons = []
            if size_diff : reasons.append("파일 크기 변경")
            if mtime_diff: reasons.append("수정 시간 변경")
            diff_reason  = " + ".join(reasons)

        if is_different:
            result['modified'].append({
                'path'       : rel,
                'reason'     : diff_reason,
                'nas_mtime'  : format_time(nas_info['mtime']),
                'local_mtime': format_time(local_info['mtime']),
                'nas_size'   : format_size(nas_info['size']),
                'local_size' : format_size(local_info['size'])
            })
        else:
            result['identical'] += 1

    return result

# ============================================================
# 💾 엑셀 저장 함수
# ============================================================
def save_to_excel(result: dict, excel_path: str):
    now        = datetime.now()
    sheet_name = now.strftime("%Y-%m-%d %H_%M")

    # 기존 파일이 있으면 열고, 없으면 새로 생성
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        base_name = sheet_name
        counter   = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base_name} ({counter})"
            counter   += 1
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    ws = wb.create_sheet(title=sheet_name)

    # 열 너비 설정
    col_widths = [6, 55, 18, 20, 18, 20, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ──────────────────────────────────────────
    # 📌 상단 정보 블록
    # ──────────────────────────────────────────
    def write_info_row(label: str, value: str, row: int) -> int:
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(bold=True, size=10, color="1F4E79")
        c.alignment = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(f"C{row}:G{row}")
        c2 = ws.cell(row=row, column=3, value=value)
        c2.font      = Font(size=10)
        c2.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 16
        return row + 1

    # 제목
    ws.merge_cells(f"A{row}:G{row}")
    title_cell = ws.cell(row=row, column=1, value="📂 NAS ↔ 로컬PC 폴더 비교 리포트  (PPT / PPTX)")
    title_cell.font      = Font(bold=True, size=14, color=COLOR["header_font"])
    title_cell.fill      = make_fill(COLOR["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    row = write_info_row("조회 일시 :"  , now.strftime("%Y년 %m월 %d일  %H:%M:%S"), row)
    row = write_info_row("NAS 경로 :"   , NAS_PATH,   row)
    row = write_info_row("로컬 경로 :"  , LOCAL_PATH, row)
    row = write_info_row("대상 확장자 :", ", ".join(sorted(TARGET_EXTENSIONS)), row)
    row = write_info_row("비교 방식 :"  , "MD5 해시" if USE_HASH_CHECK else "파일 크기 + 수정 시간", row)
    row += 1

    # ──────────────────────────────────────────
    # 📊 요약 블록
    # ──────────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    sh = ws.cell(row=row, column=1, value="📊 요약")
    sh.font      = Font(bold=True, size=11, color=COLOR["header_font"])
    sh.fill      = make_fill("2E75B6")
    sh.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    summary_data = [
        ("✅ 동일한 파일",        result['identical'],                                                          COLOR["identical"]),
        ("🆕 NAS에만 있는 파일",  len(result['only_in_nas']),                                                   COLOR["nas_only"]),
        ("📁 로컬에만 있는 파일", len(result['only_in_local']),                                                 COLOR["local_only"]),
        ("✏️  수정된 파일",        len(result['modified']),                                                      COLOR["modified"]),
        ("🔍 총 차이 항목",       len(result['only_in_nas']) + len(result['only_in_local']) + len(result['modified']), "F4B942"),
    ]

    for label, count, color in summary_data:
        ws.merge_cells(f"A{row}:D{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.font      = Font(bold=True, size=10)
        lc.fill      = make_fill(color)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        lc.border    = thin_border()

        ws.merge_cells(f"E{row}:G{row}")
        vc = ws.cell(row=row, column=5, value=f"{count:,} 개")
        vc.font      = Font(bold=True, size=10)
        vc.fill      = make_fill(color)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # ──────────────────────────────────────────
    # 공통 유틸 함수
    # ──────────────────────────────────────────
    def write_section_header(title: str, headers: list, color: str, row: int) -> int:
        ws.merge_cells(f"A{row}:G{row}")
        sc = ws.cell(row=row, column=1, value=title)
        sc.font      = Font(bold=True, size=11, color="FFFFFF")
        sc.fill      = make_fill(color)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        for col_idx, header in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_idx, value=header)
            c.font      = Font(bold=True, size=9, color="FFFFFF")
            c.fill      = make_fill(COLOR["header_bg"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border()
        ws.row_dimensions[row].height = 18
        return row + 1

    def write_no_data(row: int) -> int:
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value="— 해당 파일 없음 —")
        c.font      = Font(italic=True, color="808080", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16
        return row + 1

    def write_data_row(data: list, fill_color: str, center_cols: tuple, row: int) -> int:
        fill = make_fill(fill_color)
        for col_idx, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill      = fill
            c.font      = Font(size=9)
            c.border    = thin_border()
            c.alignment = Alignment(
                vertical  ="center",
                horizontal="center" if col_idx in center_cols else "left",
                wrap_text =(col_idx == 2)
            )
        ws.row_dimensions[row].height = 16
        return row + 1

    # ──────────────────────────────────────────
    # 🆕 섹션 1 : NAS에만 있는 파일
    # ──────────────────────────────────────────
    row = write_section_header(
        f"🆕 NAS에만 있는 파일  ({len(result['only_in_nas'])}개)  — 로컬에 없는 파일",
        ["No.", "파일 경로 (상대)", "파일 크기", "NAS 수정 시간"],
        COLOR["section_nas"], row
    )
    if result['only_in_nas']:
        for i, item in enumerate(result['only_in_nas'], 1):
            row = write_data_row(
                [i, item['path'], item['size'], item['mtime']],
                COLOR["nas_only"], (1, 3, 4), row
            )
    else:
        row = write_no_data(row)

    row += 1

    # ──────────────────────────────────────────
    # 📁 섹션 2 : 로컬에만 있는 파일
    # ──────────────────────────────────────────
    row = write_section_header(
        f"📁 로컬에만 있는 파일  ({len(result['only_in_local'])}개)  — NAS에 없는 파일",
        ["No.", "파일 경로 (상대)", "파일 크기", "로컬 수정 시간"],
        COLOR["section_loc"], row
    )
    if result['only_in_local']:
        for i, item in enumerate(result['only_in_local'], 1):
            row = write_data_row(
                [i, item['path'], item['size'], item['mtime']],
                COLOR["local_only"], (1, 3, 4), row
            )
    else:
        row = write_no_data(row)

    row += 1

    # ──────────────────────────────────────────
    # ✏️ 섹션 3 : 수정된 파일
    # ──────────────────────────────────────────
    row = write_section_header(
        f"✏️  수정된 파일  ({len(result['modified'])}개)  — 양쪽에 있지만 내용이 다른 파일",
        ["No.", "파일 경로 (상대)", "NAS 크기", "NAS 수정 시간",
         "로컬 크기", "로컬 수정 시간", "변경 사유"],
        COLOR["section_mod"], row
    )
    if result['modified']:
        for i, item in enumerate(result['modified'], 1):
            row = write_data_row(
                [i, item['path'],
                 item['nas_size'], item['nas_mtime'],
                 item['local_size'], item['local_mtime'],
                 item['reason']],
                COLOR["modified"], (1, 3, 4, 5, 6), row
            )
    else:
        row = write_no_data(row)

    # 틀 고정
    ws.freeze_panes = "A2"

    # 저장
    wb.save(excel_path)
    print(f"\n  💾 엑셀 저장 완료 : {excel_path}")
    print(f"  📋 시트 이름      : {sheet_name}")
    print(f"  📚 전체 시트 수   : {len(wb.sheetnames)}개  {wb.sheetnames}")


# ============================================================
# 🚀 메인 실행
# ============================================================
if __name__ == "__main__":
    result = compare_folders(NAS_PATH, LOCAL_PATH)

    total_diff = (len(result['only_in_nas'])
                + len(result['only_in_local'])
                + len(result['modified']))

    print("\n" + "=" * 65)
    print(f"  ✅ 동일한 파일       : {result['identical']:>6,} 개")
    print(f"  🆕 NAS에만 있는 파일 : {len(result['only_in_nas']):>6,} 개")
    print(f"  📁 로컬에만 있는 파일: {len(result['only_in_local']):>6,} 개")
    print(f"  ✏️  수정된 파일       : {len(result['modified']):>6,} 개")
    print(f"  🔍 총 차이 항목      : {total_diff:>6,} 개")
    print("=" * 65)

    save_to_excel(result, EXCEL_PATH)
